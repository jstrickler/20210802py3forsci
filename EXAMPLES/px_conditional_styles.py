#!/usr/bin/env python
import openpyxl as px
from openpyxl.formatting.rule import Rule
from openpyxl.styles import colors, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle


def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']

    color_parties(ws)

    wb.save('presidents4.xlsx')

def color_parties(ws):
    """Make Republicans red and Democrats blue"""
    r_text = Font(color="FF0000")  # red
    d_text = Font(color="0000FF")
#    red_fill = PatternFill(bgColor="FFC7CE")
    r_dxf = DifferentialStyle(font=r_text)
    d_dxf = DifferentialStyle(font=d_text)
    r_rule = Rule(type="expression", operator="beginsWith", dxf=r_dxf)
    d_rule = Rule(type="expression", operator="beginsWith", dxf=d_dxf)
    r_rule.formula = ['$J2="Republican"']
    d_rule.formula = ['$J2="Democratic"']
    ws.conditional_formatting.add('J2:J47', r_rule)
    ws.conditional_formatting.add('J2:J47', d_rule)

if __name__ == '__main__':
    main()
