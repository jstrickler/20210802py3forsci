#!/usr/bin/env python
import openpyxl as px

WORKSHEET_PATH = '../DATA/presidents.xlsx'

def main():
    """program entry point"""
    wb = px.load_workbook(WORKSHEET_PATH)
    ws = wb['US Presidents']

    print(ws.dimensions)
    print(ws.min_column)
    print(ws.min_row)
    print(ws.max_column)
    print(ws.max_row)

    # returns value for specified cell
    print(ws.cell(row=2, column=3).value, ws.cell(row=2, column=2).value, '\n')

    ws.cell(row=2, column=3).value = "Frank"
    ws['K3'].value = "HELLO"
    ws.cell(row=80, column=11).value = "IT'S ME"

    wb.save(WORKSHEET_PATH)

if __name__ == '__main__':
    main()
